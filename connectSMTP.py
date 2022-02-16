import openpyxl
import smtplib
wb = openpyxl.load_workbook('base.xlsx')
sheet = wb['Planilha1']
#percorrendo a coluna de status do pagamento e filtrando os VENCIDOS
dados_clientes = {}
for i in range(1, sheet.max_row + 1):
    pagamento = sheet.cell(row=i, column=6).value
    if pagamento == 'VENCIDO':
        cliente = sheet.cell(row=i, column=1).value
        email = sheet.cell(row=i, column=2).value
        nota = sheet.cell(row=i, column=3).value
        dados_clientes[cliente] = email,nota
    i +=1

#conectando-se a um servidor SMTP
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
#iniciando criptografia TLS
smtpObj.starttls()
#fazendo login no servidor SMTP'
login = str(input('LOGIN: '))
senha = str(input('SENHA: '))
smtpObj.login(login,senha)
print("conectado com sucesso")
#enviando um email
print('-'*30)
for cliente, (email,nota) in dados_clientes.items():
    print('-'*30)
    assunto = str(f'AVISO DE VENCIMENTO DO BOLETO N°{nota}[{cliente}]')
    mensagem = str(f'Prezado cliente,\n'
                   f'O seu boleto bancário n°{nota} está vencido!\n'
                   f'Por favor! efetue o pagamento assim que possível\n'
                   f'Agradeço desde já\n'
                   f'Atenciosamente\n'
                   f'João Gabriel\n')
    smtpObj.sendmail(login,email,f'Subject:{assunto}\n{mensagem}'.encode('utf-8'))

smtpObj.quit()
